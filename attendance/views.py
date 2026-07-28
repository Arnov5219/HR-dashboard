import csv
from datetime import date, datetime, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font as xlFont, Alignment, PatternFill, Border, Side

from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.conf import settings
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt

from attendance.services import ExcelAttendanceService

def dashboard_view(request):
    """Render the main attendance dashboard template."""
    return render(request, 'dashboard.html', {
        'today_date': date.today(),
        'office_start_time': getattr(settings, 'OFFICE_START_TIME', '09:30')
    })

def api_stats_view(request):
    """JSON view returning today's dashboard counts."""
    try:
        today = date.today()
        stats = ExcelAttendanceService.get_stats_for_date(today)
        return JsonResponse(stats)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def api_charts_view(request):
    """JSON view returning datasets for Chart.js."""
    try:
        today = date.today()
        
        # 1. Today's Distribution (Pie Chart)
        stats = ExcelAttendanceService.get_stats_for_date(today)
        pie_data = {
            'labels': ['Present', 'Late', 'Absent'],
            'datasets': [{
                'data': [stats['present'], stats['late'], stats['absent']],
                'backgroundColor': ['#0B5E8E', '#F2994A', '#EB5757']
            }]
        }
        
        # Load workbook data to build trends
        employees, records = ExcelAttendanceService.load_data()
        unique_dates = sorted(list(set(r['date'] for r in records)))
        
        # 2. Weekly Trend (Bar Chart - last 7 distinct dates in the file + today)
        # Ensure today is considered
        all_dates = list(unique_dates)
        if today not in all_dates:
            all_dates.append(today)
        all_dates = sorted(list(set(all_dates)))
        
        # Get last 7 days with data
        weekly_dates = all_dates[-7:]
        weekly_labels = [d.strftime('%a, %d %b') for d in weekly_dates]
        
        weekly_present = []
        weekly_absent = []
        
        for d in weekly_dates:
            day_stats = ExcelAttendanceService.get_stats_for_date(d)
            # Present count is total minus absent
            present_cnt = day_stats['present'] + day_stats['late']
            weekly_present.append(present_cnt)
            weekly_absent.append(day_stats['absent'])
            
        bar_data = {
            'labels': weekly_labels,
            'datasets': [
                {
                    'label': 'Present',
                    'data': weekly_present,
                    'backgroundColor': '#0B5E8E'
                },
                {
                    'label': 'Absent',
                    'data': weekly_absent,
                    'backgroundColor': '#EB5757'
                }
            ]
        }
        
        # 3. Monthly Trend (Line Chart - daily present counts for the current month/last 30 days)
        # Filter dates to current month, or default to last 30 dates if not enough
        this_month_start = today.replace(day=1)
        monthly_dates = [d for d in all_dates if d >= this_month_start]
        if len(monthly_dates) < 15:
            # Fall back to last 30 dates in general
            monthly_dates = all_dates[-30:]
            
        monthly_labels = [d.strftime('%d %b') for d in monthly_dates]
        monthly_counts = []
        for d in monthly_dates:
            day_stats = ExcelAttendanceService.get_stats_for_date(d)
            present_cnt = day_stats['present'] + day_stats['late']
            monthly_counts.append(present_cnt)
            
        line_data = {
            'labels': monthly_labels,
            'datasets': [{
                'label': 'Present Count',
                'data': monthly_counts,
                'borderColor': '#2D9CDB',
                'backgroundColor': 'rgba(45, 156, 219, 0.1)',
                'fill': True,
                'tension': 0.3
            }]
        }
        
        return JsonResponse({
            'pie': pie_data,
            'bar': bar_data,
            'line': line_data
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def api_filters_view(request):
    """JSON view returning values for excel-like filter dropdowns."""
    try:
        employees, records = ExcelAttendanceService.load_data()
        
        # Employees list
        emp_list = [{'id': emp_id, 'name': emp_name} for emp_id, emp_name in sorted(employees.items())]
        
        # Status choices
        statuses = ['Present', 'Late', 'Absent']
        
        # Date hierarchy Year -> Month -> Day
        # Group dates dynamically
        date_hierarchy = {}
        unique_dates = sorted(list(set(r['date'] for r in records)), reverse=True)
        
        # Months names mapping helper
        month_names = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
        
        for d in unique_dates:
            yr = str(d.year)
            m_idx = d.month - 1
            m_name = month_names[m_idx]
            day = d.day
            
            if yr not in date_hierarchy:
                date_hierarchy[yr] = {}
            if m_name not in date_hierarchy[yr]:
                date_hierarchy[yr][m_name] = []
            if day not in date_hierarchy[yr][m_name]:
                date_hierarchy[yr][m_name].append(day)
                
        return JsonResponse({
            'employees': emp_list,
            'statuses': statuses,
            'dates': date_hierarchy
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def _parse_request_filters(request):
    """Helper to parse query parameters for sorting, searching and filtering."""
    search_query = request.GET.get('search', '').strip()
    
    employee_ids = request.GET.get('employees', '')
    if employee_ids:
        employee_ids = [x.strip() for x in employee_ids.split(',') if x.strip()]
    else:
        employee_ids = []
        
    status_list = request.GET.get('statuses', '')
    if status_list:
        status_list = [x.strip() for x in status_list.split(',') if x.strip()]
    else:
        status_list = []
        
    # Date Range Filter
    date_range_type = request.GET.get('date_range', '')
    start_date, end_date = None, None
    today = date.today()
    
    if date_range_type == 'today':
        start_date, end_date = today, today
    elif date_range_type == 'yesterday':
        yest = today - timedelta(days=1)
        start_date, end_date = yest, yest
    elif date_range_type == 'last_7_days':
        start_date, end_date = today - timedelta(days=6), today
    elif date_range_type == 'this_month':
        start_date, end_date = today.replace(day=1), today
    elif date_range_type == 'custom':
        custom_start = request.GET.get('start_date', '')
        custom_end = request.GET.get('end_date', '')
        start_date = ExcelAttendanceService._parse_date(custom_start)
        end_date = ExcelAttendanceService._parse_date(custom_end)
    elif date_range_type == 'selected_dates':
        # Used for Excel-like multi-select hierarchical date checklist
        # Expected format: dates=2026-07-22,2026-07-23...
        selected_dates_str = request.GET.get('dates', '')
        if selected_dates_str:
            dates_list = [ExcelAttendanceService._parse_date(x.strip()) for x in selected_dates_str.split(',') if x.strip()]
            dates_list = [d for d in dates_list if d is not None]
            return search_query, employee_ids, status_list, None, None, dates_list
            
    return search_query, employee_ids, status_list, start_date, end_date, None

def _apply_custom_filters(records, search_query, employee_ids, status_list, start_date, end_date, selected_dates):
    """Helper to apply parsed filters on records."""
    # 1. Apply base filters using service
    date_range = None
    if start_date or end_date:
        date_range = {'start': start_date, 'end': end_date}
        
    filtered = ExcelAttendanceService.get_filtered_records(
        records,
        search_query=search_query,
        employee_ids=employee_ids,
        date_range=date_range,
        status_list=status_list
    )
    
    # 2. Apply multi-select hierarchical date checklist if active
    if selected_dates:
        dates_set = set(selected_dates)
        filtered = [r for r in filtered if r['date'] in dates_set]
        
    return filtered

def _paginate_and_serialize(request, records):
    """Helper to sort, paginate and serialize JSON response."""
    # Sorting parameters
    sort_col = request.GET.get('sort_col', '')
    sort_dir = request.GET.get('sort_dir', 'asc') # asc or desc
    
    # Sort
    sorted_records = ExcelAttendanceService.get_filtered_records(records, sort_col=sort_col, sort_dir=sort_dir)
    
    # Pagination parameters
    try:
        page = int(request.GET.get('page', 1))
        limit = int(request.GET.get('limit', 25))
    except ValueError:
        page = 1
        limit = 25
        
    total_count = len(sorted_records)
    
    # Calculate slices
    start = (page - 1) * limit
    end = start + limit
    paginated = sorted_records[start:end]
    
    # Serialize times & dates for JSON
    serialized = []
    for r in paginated:
        serialized.append({
            'employee_id': r['employee_id'],
            'employee_name': r['employee_name'],
            'date': r['date'].strftime('%Y-%m-%d') if r['date'] else None,
            'in_time': r['in_time'].strftime('%H:%M:%S') if r['in_time'] else None,
            'out_time': r['out_time'].strftime('%H:%M:%S') if r['out_time'] else None,
            'total_hours': r['total_hours'],
            'status': r['status']
        })
        
    import math
    pages = math.ceil(total_count / limit) if limit > 0 else 1
    
    return {
        'records': serialized,
        'total_count': total_count,
        'page': page,
        'limit': limit,
        'pages': pages
    }

def api_attendance_today_view(request):
    """JSON endpoint for Today's Attendance records."""
    try:
        today = date.today()
        # Today's records are processed (with synthethic Absent records for employees not present)
        records = ExcelAttendanceService.get_processed_records(today)
        
        # Parse and apply search/filters
        search_q, emp_ids, statuses, start_d, end_d, sel_dates = _parse_request_filters(request)
        filtered = _apply_custom_filters(records, search_q, emp_ids, statuses, start_d, end_d, sel_dates)
        
        # Paginate & return
        data = _paginate_and_serialize(request, filtered)
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def api_attendance_history_view(request):
    """JSON endpoint for Historical Attendance records."""
    try:
        today = date.today()
        records = ExcelAttendanceService.get_all_attendance_history(today)
        
        # Parse and apply search/filters
        search_q, emp_ids, statuses, start_d, end_d, sel_dates = _parse_request_filters(request)
        filtered = _apply_custom_filters(records, search_q, emp_ids, statuses, start_d, end_d, sel_dates)
        
        # Paginate & return
        data = _paginate_and_serialize(request, filtered)
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def api_export_view(request):
    """Export current filtered view to CSV or Excel formats."""
    try:
        export_type = request.GET.get('type', 'today') # today or history
        export_format = request.GET.get('format', 'csv') # csv or excel
        
        today = date.today()
        if export_type == 'today':
            records = ExcelAttendanceService.get_processed_records(today)
        else:
            records = ExcelAttendanceService.get_all_attendance_history(today)
            
        # Parse and apply search/filters
        search_q, emp_ids, statuses, start_d, end_d, sel_dates = _parse_request_filters(request)
        filtered = _apply_custom_filters(records, search_q, emp_ids, statuses, start_d, end_d, sel_dates)
        
        # Apply sorting before export
        sort_col = request.GET.get('sort_col', '')
        sort_dir = request.GET.get('sort_dir', 'asc')
        filtered = ExcelAttendanceService.get_filtered_records(filtered, sort_col=sort_col, sort_dir=sort_dir)
        
        filename = f"attendance_{export_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        if export_format == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
            
            writer = csv.writer(response)
            writer.writerow(['Employee ID', 'Employee Name', 'Date', 'IN', 'OUT', 'Total Hours', 'Status'])
            for r in filtered:
                writer.writerow([
                    r['employee_id'],
                    r['employee_name'],
                    r['date'].strftime('%Y-%m-%d') if r['date'] else '',
                    r['in_time'].strftime('%H:%M:%S') if r['in_time'] else '-',
                    r['out_time'].strftime('%H:%M:%S') if r['out_time'] else '-',
                    r['total_hours'] or '-',
                    r['status']
                ])
            return response
            
        elif export_format == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Export"
            
            # Styles
            header_font = xlFont(name="Arial", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0B5E8E", end_color="0B5E8E", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            thin_border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
            
            # Headers
            headers = ['Employee ID', 'Employee Name', 'Date', 'IN', 'OUT', 'Total Hours', 'Status']
            ws.append(headers)
            
            # Format headers
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                
            # Values
            for r in filtered:
                row_val = [
                    r['employee_id'],
                    r['employee_name'],
                    r['date'].strftime('%Y-%m-%d') if r['date'] else '',
                    r['in_time'].strftime('%H:%M:%S') if r['in_time'] else '-',
                    r['out_time'].strftime('%H:%M:%S') if r['out_time'] else '-',
                    r['total_hours'] or '-',
                    r['status']
                ]
                ws.append(row_val)
                
            # Apply borders and auto column width
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=7):
                for cell in row:
                    cell.border = thin_border
                    if cell.row > 1 and cell.column in (3, 4, 5, 6, 7):
                        cell.alignment = Alignment(horizontal="center")
                        
            # Adjust column width
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
            # Save to stream
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
            return response
            
    except Exception as e:
        return HttpResponse(f"Export failed: {str(e)}", status=500)
