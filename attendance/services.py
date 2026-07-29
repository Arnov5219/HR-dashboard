import os
import threading
from datetime import date, datetime, time, timedelta
import openpyxl
from django.conf import settings

class ExcelAttendanceService:
    _lock = threading.Lock()
    _cached_data = None
    _cached_mtime = None
    _cached_employees = None

    @classmethod
    def get_excel_path(cls):
        # Fallback to local copy if the absolute configured path doesn't exist
        excel_path = getattr(settings, 'ATTENDANCE_EXCEL_PATH', None)
        if not excel_path or not os.path.exists(excel_path):
            # Try workspace root
            excel_path = os.path.join(settings.BASE_DIR, 'HR_Master.xlsm')
        return excel_path

    @classmethod
    def load_data(cls):
        excel_path = cls.get_excel_path()
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel file not found at path: {excel_path}")
        
        mtime = os.path.getmtime(excel_path)
        
        with cls._lock:
            if cls._cached_data is not None and cls._cached_mtime == mtime:
                return cls._cached_employees, cls._cached_data
            
            # Load workbook
            # data_only=True evaluates Excel formulas (like TOTAL HOURS formulas)
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            
            try:
                # 1. Parse CONFIG sheet for master employee list
                employees = {}
                if 'CONFIG' in wb.sheetnames:
                    config_sheet = wb['CONFIG']
                    # Expecting headers in row 1: 'Employee ID', 'Employee Name'
                    for row in config_sheet.iter_rows(min_row=2, values_only=True):
                        if not row or len(row) < 2:
                            continue
                        emp_id_val, emp_name_val = row[0], row[1]
                        if emp_id_val is not None:
                            emp_id = str(emp_id_val).strip()
                            emp_name = str(emp_name_val).strip() if emp_name_val else ""
                            employees[emp_id] = emp_name
                
                # 2. Parse Attendance sheet
                records = []
                if 'Attendance' in wb.sheetnames:
                    sheet = wb['Attendance']
                    first_row = next(sheet.iter_rows(max_row=1), None)
                    if first_row:
                        headers = [str(c.value).strip().lower() if c.value is not None else "" for c in first_row]
                        
                        # Map indices
                        # Columns are: Employee ID, Employee Name, Date, IN, OUT, TOTAL HOURS
                        col_map = {
                            'employee_id': 0,
                            'employee_name': 1,
                            'date': 2,
                            'in_time': 3,
                            'out_time': 4,
                            'total_hours': 5
                        }
                        for key, header_name in [('employee_id', 'employee id'), ('employee_name', 'employee name'), ('date', 'date'), ('in_time', 'in'), ('out_time', 'out'), ('total_hours', 'total hours')]:
                            if header_name in headers:
                                col_map[key] = headers.index(header_name)
                        
                        for row in sheet.iter_rows(min_row=2, values_only=True):
                            if not row or len(row) <= max(col_map.values()):
                                continue
                            
                            emp_id_val = row[col_map['employee_id']]
                            if emp_id_val is None or str(emp_id_val).strip() == "":
                                continue
                            
                            emp_id = str(emp_id_val).strip()
                            emp_name = str(row[col_map['employee_name']]).strip() if row[col_map['employee_name']] is not None else ""
                            
                            # Parse date
                            raw_date = row[col_map['date']]
                            parsed_date = cls._parse_date(raw_date)
                            if not parsed_date:
                                continue
                            
                            # Parse IN/OUT times
                            in_time = cls._parse_time(row[col_map['in_time']])
                            out_time = cls._parse_time(row[col_map['out_time']])
                            
                            # Parse or compute TOTAL HOURS
                            total_hours = cls._parse_total_hours(row[col_map['total_hours']], in_time, out_time)
                            
                            records.append({
                                'employee_id': emp_id,
                                'employee_name': emp_name,
                                'date': parsed_date,
                                'in_time': in_time,
                                'out_time': out_time,
                                'total_hours': total_hours
                            })
                            
                            # If employee is not in master list from CONFIG, add them dynamically
                            if emp_id not in employees:
                                employees[emp_id] = emp_name

                cls._cached_data = records
                cls._cached_employees = employees
                cls._cached_mtime = mtime
                return employees, records
            finally:
                wb.close()

    @classmethod
    def _parse_date(cls, val):
        if val is None:
            return None
        if isinstance(val, (datetime, date)):
            return val.date() if isinstance(val, datetime) else val
        if isinstance(val, str):
            clean_val = val.strip().split()[0] if val.strip() else ""
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
                try:
                    return datetime.strptime(clean_val, fmt).date()
                except ValueError:
                    continue
        return None

    @classmethod
    def _parse_time(cls, val):
        if val is None or val == "":
            return None
        if isinstance(val, time):
            return val
        if isinstance(val, datetime):
            return val.time()
        if isinstance(val, (int, float)):
            total_seconds = int(val * 86400)
            hours = min(max(total_seconds // 3600, 0), 23)
            minutes = min(max((total_seconds % 3600) // 60, 0), 59)
            seconds = min(max(total_seconds % 60, 0), 59)
            return time(hours, minutes, seconds)
        if isinstance(val, str):
            clean_val = val.strip()
            for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M:%S %p"):
                try:
                    return datetime.strptime(clean_val, fmt).time()
                except ValueError:
                    continue
        return None

    @classmethod
    def _parse_total_hours(cls, val, in_time, out_time):
        if val is not None and val != "":
            if isinstance(val, timedelta):
                total_seconds = int(val.total_seconds())
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                return f"{hours:02d}:{minutes:02d}"
            if isinstance(val, time):
                return val.strftime("%H:%M")
            if isinstance(val, datetime):
                return val.strftime("%H:%M")
            if isinstance(val, (int, float)):
                total_hours = val * 24
                hours = int(total_hours)
                minutes = int((total_hours - hours) * 60 + 0.5)
                return f"{hours:02d}:{minutes:02d}"
            return str(val).strip()
        
        # Calculate as fallback
        if in_time and out_time:
            today = date.today()
            dt_in = datetime.combine(today, in_time)
            dt_out = datetime.combine(today, out_time)
            if dt_out >= dt_in:
                diff = dt_out - dt_in
                total_seconds = int(diff.total_seconds())
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                return f"{hours:02d}:{minutes:02d}"
        return None

    @classmethod
    def get_processed_records(cls, target_date):
        employees, records = cls.load_data()
        
        # Filter records for target date
        date_records = [r for r in records if r['date'] == target_date]
        date_records_map = {r['employee_id']: r for r in date_records}
        
        processed = []
        for emp_id, emp_name in employees.items():
            if emp_id in date_records_map:
                rec = dict(date_records_map[emp_id])
                processed.append(rec)
            else:
                # Synthesize Absent record
                processed.append({
                    'employee_id': emp_id,
                    'employee_name': emp_name,
                    'date': target_date,
                    'in_time': None,
                    'out_time': None,
                    'total_hours': None
                })
        return processed

    @classmethod
    def get_all_attendance_history(cls, today_date):
        _, records = cls.load_data()
        
        # Return all records (including today's)
        history_records = list(records)
        
        processed = []
        for r in history_records:
            rec = dict(r)
            processed.append(rec)
            
        return processed

    @classmethod
    def get_filtered_records(cls, records, search_query=None, employee_ids=None, date_range=None, sort_col=None, sort_dir=None, weeks=None):
        filtered = list(records)
        
        # 1. Global Search Query
        if search_query:
            q = str(search_query).lower()
            filtered = [
                r for r in filtered
                if q in r['employee_id'].lower() or
                   q in r['employee_name'].lower() or
                   (r['date'] and q in r['date'].strftime('%Y-%m-%d'))
            ]
            
        # 2. Employee IDs filter (Multi-select)
        if employee_ids:
            if isinstance(employee_ids, str):
                employee_ids = [x.strip() for x in employee_ids.split(',') if x.strip()]
            employee_ids_set = set(employee_ids)
            filtered = [r for r in filtered if r['employee_id'] in employee_ids_set]
            
        # 3. Date Range filter
        if date_range:
            start_date, end_date = None, None
            if isinstance(date_range, dict):
                start_date = date_range.get('start')
                end_date = date_range.get('end')
            elif isinstance(date_range, (list, tuple)) and len(date_range) == 2:
                start_date, end_date = date_range[0], date_range[1]
                
            if start_date:
                filtered = [r for r in filtered if r['date'] >= start_date]
            if end_date:
                filtered = [r for r in filtered if r['date'] <= end_date]
                
        # 4. Weeks filter (Multi-select)
        if weeks:
            if isinstance(weeks, str):
                weeks = [x.strip() for x in weeks.split(',') if x.strip()]
            weeks_set = set(weeks)
            filtered = [r for r in filtered if r['date'] and f"Week {r['date'].isocalendar().week} ({r['date'].year})" in weeks_set]
            
        # 5. Sorting
        if sort_col:
            key_map = {
                'employee': 'employee_name',
                'employee_name': 'employee_name',
                'employee_id': 'employee_id',
                'date': 'date',
                'in': 'in_time',
                'in_time': 'in_time',
                'out': 'out_time',
                'out_time': 'out_time',
                'total_hours': 'total_hours'
            }
            sort_key = key_map.get(sort_col.lower(), sort_col)
            
            def make_sort_key(item):
                val = item.get(sort_key)
                if val is None:
                    if sort_key in ('in_time', 'out_time'):
                        return time(23, 59, 59)
                    if sort_key == 'date':
                        return date.min
                    if sort_key == 'total_hours':
                        return ""
                    return ""
                return val

            reverse_sort = (sort_dir == 'desc')
            filtered.sort(key=make_sort_key, reverse=reverse_sort)
            
        return filtered
