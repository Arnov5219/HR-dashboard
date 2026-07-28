"""Module to handle reading attendance records from HR_Master.xlsm Excel file.

This module parses the new sheet structure:
A - Employee ID
B - Employee Name
C - Date
D - IN
E - OUT
F - Last Punch
G - Status
H - Last Updated
"""

import os
from datetime import datetime, date, time
import openpyxl
from django.conf import settings

def parse_excel_date(val):
    """Safely parse date values from Excel cell."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    if isinstance(val, str):
        # Extract the date part (before spaces) in case it's a datetime string
        clean_val = val.strip().split()[0] if val.strip() else ""
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(clean_val, fmt).date()
            except ValueError:
                continue
    return None

def parse_excel_time(val):
    """Safely parse time values from Excel cell, handling time objects, datetimes, strings, and numeric fractions."""
    if val is None or val == "":
        return None
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, (int, float)):
        # Excel stores time as a fraction of a 24-hour day
        total_seconds = int(val * 86400)
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        # Bound variables to valid time range
        hours = min(max(hours, 0), 23)
        minutes = min(max(minutes, 0), 59)
        seconds = min(max(seconds, 0), 59)
        return time(hours, minutes, seconds)
    if isinstance(val, str):
        clean_val = val.strip()
        for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M:%S %p"):
            try:
                return datetime.strptime(clean_val, fmt).time()
            except ValueError:
                continue
    return None

def parse_excel_datetime(val):
    """Safely parse datetime/last_updated values from Excel cell."""
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime.combine(val, datetime.min.time())
    if isinstance(val, str):
        clean_val = val.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d-%m-%Y %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%m/%d/%Y %I:%M:%S %p"):
            try:
                return datetime.strptime(clean_val, fmt)
            except ValueError:
                continue
    return str(val)

def parse_excel_total_hours(cell):
    """
    Safely parse total hours from an Excel cell.
    Handles datetime, time, float/int and string, formatting it as a string
    matching the Excel display (e.g. HH:MM or decimal format).
    """
    if cell is None:
        return None
    val = cell.value
    if val is None or val == "":
        return None
    
    num_fmt = cell.number_format or ""
    
    if isinstance(val, datetime):
        return val.strftime("%H:%M")
    if isinstance(val, time):
        return val.strftime("%H:%M")
    if isinstance(val, (int, float)):
        # If number format doesn't look like time (e.g., standard decimal format)
        if "h" not in num_fmt.lower() and "m" not in num_fmt.lower():
            return f"{val:.2f}"
        
        # Excel stores time as a fraction of a 24-hour day
        total_hours = val * 24
        wrapped_hours = total_hours % 24
        hours = int(wrapped_hours)
        minutes = int((wrapped_hours - hours) * 60 + 0.5)
        if minutes >= 60:
            hours += 1
            minutes -= 60
        hours = hours % 24
        return f"{hours:02d}:{minutes:02d}"
    if isinstance(val, str):
        return val.strip()
    return str(val)


def read_attendance_data():
    """
    Reads attendance records from HR_Master.xlsm.
    Returns a list of dictionaries containing raw attendance data.
    Raises FileNotFoundError if the file is missing.
    Raises PermissionError if the file is locked or inaccessible.
    """
    excel_path = getattr(settings, 'ATTENDANCE_EXCEL_PATH', None)
    if not excel_path:
        raise FileNotFoundError("Excel path configuration missing in Django settings.")

    # Check existence
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found at path: {excel_path}")

    try:
        # Load workbook with keep_vba=True and data_only=True
        wb = openpyxl.load_workbook(excel_path, keep_vba=True, data_only=True)
    except PermissionError as e:
        raise PermissionError(f"Permission denied. The Excel file may be locked or open in another application.") from e
    except Exception as e:
        raise IOError(f"Failed to read the Excel file: {str(e)}") from e

    try:
        sheet = wb.active
        if not sheet:
            return []

        # Read first row as headers
        first_row = next(sheet.iter_rows(max_row=1), None)
        if not first_row:
            return []

        headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in first_row]

        # Map expected headers to their column indices
        col_mappings = {
            'employee_id': ['employee id', 'employee_id', 'id', 'emp id', 'empid'],
            'employee_name': ['employee name', 'employee_name', 'name', 'emp name', 'empname'],
            'date': ['date'],
            'in_time': ['in', 'in time', 'in_time'],
            'out_time': ['out', 'out time', 'out_time'],
            'last_punch': ['last punch', 'last_punch'],
            'status': ['status'],
            'last_updated': ['last updated', 'last_updated'],
            'total_hours': ['total hours', 'total_hours', 'total', 'hours']
        }

        indices = {}
        for key, variants in col_mappings.items():
            found_idx = None
            for variant in variants:
                if variant in headers:
                    found_idx = headers.index(variant)
                    break
            indices[key] = found_idx

        # If headers matching fails, fall back to exact column offsets (A to I)
        if indices['employee_id'] is None:
            indices['employee_id'] = 0
        if indices['employee_name'] is None:
            indices['employee_name'] = 1
        if indices['date'] is None:
            indices['date'] = 2
        if indices['in_time'] is None:
            indices['in_time'] = 3
        if indices['out_time'] is None:
            indices['out_time'] = 4
        if indices['last_punch'] is None:
            indices['last_punch'] = 5
        if indices['status'] is None:
            indices['status'] = 6
        if indices['last_updated'] is None:
            indices['last_updated'] = 7
        if indices.get('total_hours') is None:
            indices['total_hours'] = 8

        attendance_records = []

        # Iterate over data rows
        for row in sheet.iter_rows(min_row=2, values_only=False):
            # Check maximum sheet length
            if len(row) <= max(indices.values()):
                continue

            emp_id_val = row[indices['employee_id']].value
            if emp_id_val is None or str(emp_id_val).strip() == "":
                continue

            emp_id = str(emp_id_val).strip()
            emp_name = str(row[indices['employee_name']].value or "").strip()
            
            # Parse Date
            raw_date = row[indices['date']].value
            rec_date = parse_excel_date(raw_date)
            if not rec_date:
                # Fallback: try parsing from last_updated if available
                last_u_val = row[indices['last_updated']].value if indices['last_updated'] is not None else None
                if last_u_val:
                    rec_date = parse_excel_date(last_u_val)
                # Fallback to today's date if still missing
                if not rec_date:
                    rec_date = date.today()

            # Parse Punch Times and Status
            in_t = parse_excel_time(row[indices['in_time']].value)
            out_t = parse_excel_time(row[indices['out_time']].value)
            last_p = parse_excel_time(row[indices['last_punch']].value)
            status = str(row[indices['status']].value or "Absent").strip()
            last_u = parse_excel_datetime(row[indices['last_updated']].value)
            total_h = parse_excel_total_hours(row[indices['total_hours']])

            # Save as a direct python dict
            attendance_records.append({
                'employee_id': emp_id,
                'employee_name': emp_name,
                'date': rec_date,
                'in_time': in_t,
                'out_time': out_t,
                'last_punch': last_p,
                'status': status,
                'last_updated': last_u,
                'total_hours': total_h
            })

        return attendance_records

    finally:
        wb.close()
